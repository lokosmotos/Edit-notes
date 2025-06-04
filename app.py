from flask import Flask, request, render_template, send_from_directory
from docx import Document
import re
import os
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'docx'}
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB limit

# Create uploads directory if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    """Check if the file has an allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def parse_docx(file_path):
    """Parse a .docx file into metadata, edits, and images."""
    doc = Document(file_path)
    data = {
        "metadata": {
            "title": "",
            "genre": "",
            "version": "",
            "language": "",
            "secondary": "",
            "subtitles": "",
            "runtime": "",
            "date": "",
            "remarks": "None",
            "filename": os.path.basename(file_path),
            "parsed_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        },
        "edits": [],
        "images": []
    }

    current_section = "metadata"
    last_metadata_key = None

    for para in doc.paragraphs:
        # Normalize text: replace tabs and multiple spaces with a single space
        text = re.sub(r'\s+', ' ', para.text).strip()
        if not text:
            continue

        # Section detection
        if text.lower() == "edit notes":
            current_section = "metadata"
            continue
        elif text.lower() == "remarks:":
            current_section = "remarks"
            continue
        elif re.match(r"^\d+\.\s*\d{2}:\d{2}:\d{2}", text):
            current_section = "edits"
            continue
        elif text.lower().endswith((".jpg", ".jpeg", ".png", ".gif")):
            current_section = "images"

        # Parse metadata
        if current_section == "metadata":
            if text in ["Title", "Genre", "Version", "Language", "Secondary", "Subtitles", "Runtime", "Date"]:
                last_metadata_key = text.lower()
            elif last_metadata_key and text:
                if ": " in text:  # Handle "Version: Complete"
                    key, value = map(str.strip, text.split(": ", 1))
                    if key.lower() in data["metadata"]:
                        data["metadata"][key.lower()] = value
                else:
                    data["metadata"][last_metadata_key] = text
                    last_metadata_key = None

        # Parse remarks
        elif current_section == "remarks":
            if text != "Remarks:":
                data["metadata"]["remarks"] = text or "None"

        # Parse edits
        elif current_section == "edits":
            match = re.match(r"(\d+)\.\s*(\d{2}:\d{2}:\d{2}(?:\s*-\s*\d{2}:\d{2}:\d{2})?)\s*(.+)", text)
            if match:
                edit_number, timestamp, description = match.groups()
                data["edits"].append({
                    "number": int(edit_number),
                    "time": timestamp,
                    "description": description.strip()
                })

        # Parse images
        elif current_section == "images":
            if text.lower().endswith((".jpg", ".jpeg", ".png", ".gif")):
                data["images"].append(text)

    return data

def create_excel(data):
    """Create an Excel file from the parsed data."""
    wb = Workbook()
    
    # Metadata Sheet
    ws_metadata = wb.active
    ws_metadata.title = "Metadata"
    ws_metadata['A1'] = "Field"
    ws_metadata['B1'] = "Value"
    ws_metadata['A1'].font = Font(bold=True)
    ws_metadata['B1'].font = Font(bold=True)
    ws_metadata.column_dimensions['A'].width = 20
    ws_metadata.column_dimensions['B'].width = 50

    metadata_rows = [
        ("Title", data["metadata"]["title"]),
        ("Genre", data["metadata"]["genre"]),
        ("Version", data["metadata"]["version"]),
        ("Primary Language", data["metadata"]["language"]),
        ("Secondary Language", data["metadata"]["secondary"]),
        ("Subtitles", data["metadata"]["subtitles"]),
        ("Runtime", data["metadata"]["runtime"]),
        ("Date", data["metadata"]["date"]),
        ("Remarks", data["metadata"]["remarks"]),
        ("Filename", data["metadata"]["filename"]),
        ("Parsed At", data["metadata"]["parsed_date"])
    ]
    for idx, (field, value) in enumerate(metadata_rows, start=2):
        ws_metadata[f'A{idx}'] = field
        ws_metadata[f'B{idx}'] = value or "N/A"
        ws_metadata[f'A{idx}'].alignment = Alignment(vertical='top')
        ws_metadata[f'B{idx}'].alignment = Alignment(vertical='top', wrap_text=True)

    # Edits Sheet
    ws_edits = wb.create_sheet(title="Edits")
    ws_edits['A1'] = "Edit Number"
    ws_edits['B1'] = "Timestamp"
    ws_edits['C1'] = "Description"
    ws_edits['A1'].font = Font(bold=True)
    ws_edits['B1'].font = Font(bold=True)
    ws_edits['C1'].font = Font(bold=True)
    ws_edits.column_dimensions['A'].width = 15
    ws_edits.column_dimensions['B'].width = 20
    ws_edits.column_dimensions['C'].width = 60

    for idx, edit in enumerate(data["edits"], start=2):
        ws_edits[f'A{idx}'] = edit["number"]
        ws_edits[f'B{idx}'] = edit["time"]
        ws_edits[f'C{idx}'] = edit["description"]
        ws_edits[f'C{idx}'].alignment = Alignment(wrap_text=True)

    # Images Sheet
    if data["images"]:
        ws_images = wb.create_sheet(title="Images")
        ws_images['A1'] = "Image Filename"
        ws_images['A1'].font = Font(bold=True)
        ws_images.column_dimensions['A'].width = 30
        for idx, image in enumerate(data["images"], start=2):
            ws_images[f'A{idx}'] = image

    # Save the workbook
    filename = secure_filename(f"{data['metadata']['filename'].rsplit('.', 1)[0]}_parsed_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    wb.save(filepath)
    return filename

@app.route('/', methods=['GET', 'POST'])
def index():
    """Handle file uploads and render the main page."""
    if request.method == 'POST':
        if 'file' not in request.files:
            return render_template('index.html', error='No file selected')
        
        file = request.files['file']
        if file.filename == '':
            return render_template('index.html', error='No file selected')
        
        if not allowed_file(file.filename):
            return render_template('index.html', error='Only .docx files are allowed')
        
        try:
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            data = parse_docx(filepath)
            excel_filename = create_excel(data)
            return render_template('index.html', download_url=f'/download/{excel_filename}')
            
        except Exception as e:
            return render_template('index.html', error=f'Error processing file: {str(e)}')
    
    return render_template('index.html', download_url=None)

@app.route('/download/<filename>')
def download_file(filename):
    """Serve a file for download."""
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=True)
